import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill

class DeliverySystemApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📦 产品发货计划交互系统 (本地绿色版)")
        self.root.geometry("700x650")
        self.root.configure(bg="#f5f7fa")
        
        # 绑定的 Excel 数据文件名
        self.filename = "含发货日期表格.xlsx"
        self.products = []
        self.date_cols = []
        
        # 样式配置
        self.setup_styles()
        
        # 初始化数据加载
        self.load_excel_base_data()
        
        # 构建前端操作页面
        self.build_ui()
        
        # 启动时自动执行一次全表检测与今日日期所需量标红
        self.auto_highlight_today_requirements()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TLabel', font=('微软雅黑', 10), background='#f5f7fa')
        style.configure('TButton', font=('微软雅黑', 10, 'bold'), padding=6)
        style.configure('TEntry', font=('微软雅黑', 10))
        style.configure('TCombobox', font=('微软雅黑', 10))

    def load_excel_base_data(self):
        """从本地同级目录下的 Excel 中动态提取产品和发货时间轴"""
        if not os.path.exists(self.filename):
            messagebox.showerror("核心文件缺失", f"未在当前目录下找到【{self.filename}】！\n请确保该 Excel 表格与本软件放在同一个文件夹里。")
            sys.exit()
        try:
            wb = openpyxl.load_workbook(self.filename, data_only=True)
            ws = wb["产品发货计划表"]
            headers = [cell.value for cell in ws[1]]
            
            # 动态筛选所有含“发货”字样的日期列
            self.date_cols = [str(h).strip() for h in headers if h and "发货" in str(h)]
            
            # 缓存基础产品库
            self.products = []
            for row in range(2, ws.max_row + 1):
                p_id = ws.cell(row=row, column=1).value
                p_name = ws.cell(row=row, column=2).value
                p_price = ws.cell(row=row, column=3).value
                if p_id:
                    self.products.append({
                        "id": str(p_id).strip(),
                        "name": str(p_name).strip(),
                        "price": float(p_price or 0.0)
                    })
            wb.close()
        except Exception as e:
            messagebox.showerror("数据读取失败", f"读取 Excel 失败，请检查工作表名称是否为 '产品发货计划表'\n错误信息: {e}")

    def build_ui(self):
        # 顶部标题栏
        title_frame = tk.Frame(self.root, bg="#1e3d59", height=60)
        title_frame.pack(fill="x", side="top")
        lbl_title = tk.Label(title_frame, text="📦 产品发货计划管理系统", font=("微软雅黑", 16, "bold"), fg="white", bg="#1e3d59")
        lbl_title.pack(pady=15)

        # 主交互区域
        main_frame = tk.Frame(self.root, padx=25, pady=15, bg="#f5f7fa")
        main_frame.pack(fill="both", expand=True)

        # ---- 1. 产品动态匹配检索区 ----
        lbl_sec1 = tk.Label(main_frame, text="一、产品检索与确认", font=("微软雅黑", 11, "bold"), fg="#1e3d59")
        lbl_sec1.pack(anchor="w", pady=(5, 5))
        
        search_frame = tk.Frame(main_frame, bg="#f5f7fa")
        search_frame.pack(fill="x", pady=5)
        
        ttk.Label(search_frame, text="输入产品编号(后2/3位或任意数):").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.execute_auto_match)
        self.ent_search = ttk.Entry(search_frame, textvariable=self.search_var, width=25)
        self.ent_search.pack(side="left", padx=10)
        
        match_frame = tk.Frame(main_frame, bg="#f5f7fa")
        match_frame.pack(fill="x", pady=5)
        ttk.Label(match_frame, text="自动匹配结果(请点击下拉确认):").pack(side="left")
        self.cb_products = ttk.Combobox(match_frame, state="readonly", width=40)
        self.cb_products.pack(side="left", padx=10)

        # ---- 2. 日期时间段管理区 ----
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=15)
        lbl_sec2 = tk.Label(main_frame, text="二、生成/扩展新日期列（非必选）", font=("微软雅黑", 11, "bold"), fg="#1e3d59")
        lbl_sec2.pack(anchor="w", pady=(0, 5))
        
        date_set_frame = tk.Frame(main_frame, bg="#f5f7fa")
        date_set_frame.pack(fill="x", pady=5)
        ttk.Label(date_set_frame, text="输入新日期 (格式如 05-20):").pack(side="left")
        self.ent_new_date = ttk.Entry(date_set_frame, width=15)
        self.ent_new_date.pack(side="left", padx=10)
        btn_add_date = ttk.Button(date_set_frame, text="➕ 依次向后追加该日期列", command=self.add_new_date_column)
        btn_add_date.pack(side="left", padx=5)

        # ---- 3. 发货需求填报区 ----
        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=15)
        lbl_sec3 = tk.Label(main_frame, text="三、需求量填报（联动乘法公式计算）", font=("微软雅黑", 11, "bold"), fg="#1e3d59")
        lbl_sec3.pack(anchor="w", pady=(0, 5))
        
        input_frame = tk.Frame(main_frame, bg="#f5f7fa")
        input_frame.pack(fill="x", pady=5)
        
        ttk.Label(input_frame, text="选择目标发货日期:").grid(row=0, column=0, sticky="w", pady=5)
        self.cb_target_date = ttk.Combobox(input_frame, state="readonly", values=self.date_cols, width=20)
        self.cb_target_date.grid(row=0, column=1, padx=10, pady=5)
        if self.date_cols: self.cb_target_date.current(0)
        
        ttk.Label(input_frame, text="输入要写入的数量:").grid(row=1, column=0, sticky="w", pady=5)
        self.ent_qty = ttk.Entry(input_frame, width=22)
        self.ent_qty.insert(0, "100")
        self.ent_qty.grid(row=1, column=1, padx=10, pady=5)

        # ---- 4. 底部核心执行按钮与状态提示 ----
        self.btn_save = tk.Button(
            self.root, 
            text="💾 确认更新并保存至 Excel (联动计算 + 自动变红检测)", 
            font=("微软雅黑", 12, "bold"), 
            bg="#17b978", 
            fg="white", 
            relief="flat",
            command=self.execute_write_and_calculate
        )
        self.btn_save.pack(fill="x", padx=40, pady=10)

        # 实时通知底栏
        self.status_bar = tk.Label(self.root, text="系统就绪。建议在修改前关闭正在查看的 Excel 表格以免冲突。", bd=1, relief="sunken", anchor="w", bg="#e4f9f5", font=("微软雅黑", 9))
        self.status_bar.pack(side="bottom", fill="x")

    def execute_auto_match(self, *args):
        """核心功能1：输入后2/3位时自动模糊匹配相关产品项"""
        search_text = self.search_var.get().strip()
        if len(search_text) >= 2:
            matched = [p for p in self.products if p["id"].endswith(search_text) or search_text in p["id"]]
            if matched:
                options = [f"{p['id']} | {p['name']} (单价: {p['price']})" for p in matched]
                self.cb_products['values'] = options
                self.cb_products.current(0)
                self.status_bar.config(text=f"🎯 自动匹配成功：找到 {len(matched)} 个相关项。")
            else:
                self.cb_products['values'] = ["❌ 未找到任何匹配结果"]
                self.cb_products.set("❌ 未找到任何匹配结果")
        else:
            self.cb_products['values'] = []
            self.cb_products.set("等待继续输入位数字...")

    def add_new_date_column(self):
        """核心功能2：用户可以追加输入发货日期段，在表格中按照日期依次向后排列"""
        raw_date = self.ent_new_date.get().strip()
        if not raw_date:
            messagebox.showwarning("提示", "请输入日期，例如：05-20")
            return
        new_col_name = f"{raw_date}发货"
        if new_col_name in self.date_cols:
            messagebox.showwarning("提示", f"日期列【{new_col_name}】已存在，无需重复添加。")
            return
            
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb["产品发货计划表"]
            headers = [cell.value for cell in ws[1]]
            
            # 定位插入位置：在“开票数量”列之前插入，以保证按照日期依次向后排列的完整度
            if "开票数量" in headers:
                insert_idx = headers.index("开票数量") + 1
            else:
                insert_idx = ws.max_column + 1
                
            ws.insert_cols(insert_idx)
            ws.cell(row=1, column=insert_idx, value=new_col_name)
            
            # 默认为已有所有行刷上初始值0
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value:
                    ws.cell(row=r, column=insert_idx, value=0.0)
                    
            wb.save(self.filename)
            wb.close()
            
            # 刷新本地缓存
            self.load_excel_base_data()
            self.cb_target_date['values'] = self.date_cols
            self.cb_target_date.set(new_col_name)
            messagebox.showinfo("成功", f"时间轴已成功延展！新日期列【{new_col_name}】已按顺序依次向后排列。")
        except Exception as e:
            messagebox.showerror("写入失败", f"无法更新日期轴，请确认 Excel 未被外部打开占满。\n错误: {e}")

    def execute_write_and_calculate(self):
        """核心功能3, 4, 5：输入量、存回首页原表、含税单价=单价*总需求量联动自动算账"""
        selected_prod = self.cb_products.get()
        target_date = self.cb_target_date.get()
        qty_str = self.ent_qty.get().strip()

        if " | " not in selected_prod:
            messagebox.showwarning("填报提示", "请先通过输入编号，并在下拉列表中选择确认产品项！")
            return
        if not target_date or "发货" not in target_date:
            messagebox.showwarning("填报提示", "请选择正确的目标发货日期！")
            return
        try:
            qty = float(qty_str)
        except ValueError:
            messagebox.showwarning("填报提示", "请输入有效的数字需求量！")
            return

        p_id = selected_prod.split(" | ")[0].strip()

        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb["产品发货计划表"]
            headers = [str(cell.value).strip() for cell in ws[1]]
            
            # 校验核心公式列名是否存在
            if "开票数量" not in headers or "含税单价" not in headers:
                messagebox.showerror("结构错误", "Excel 表头缺少 '开票数量' 或 '含税单价' 列，请检查表格原始模板。")
                return
                
            qty_sum_col_idx = headers.index("开票数量") + 1
            tax_price_col_idx = headers.index("含税单价") + 1
            target_date_col_idx = headers.index(target_date) + 1
            
            # 定位行
            target_row_idx = None
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value).strip() == p_id:
                    target_row_idx = row
                    break
            
            if target_row_idx:
                # 3. 填入指定日期数量
                ws.cell(row=target_row_idx, column=target_date_col_idx, value=qty)
                
                # 4. 动态计算开票数量总和 (所有日期列的数据相加)
                date_indices = [headers.index(d) + 1 for d in self.date_cols if d in headers]
                total_qty_sum = 0.0
                for c_idx in date_indices:
                    val = ws.cell(row=target_row_idx, column=c_idx).value
                    total_qty_sum += float(val or 0.0)
                
                ws.cell(row=target_row_idx, column=qty_sum_col_idx, value=total_qty_sum)
                
                # 4. 联动含税单价公式：单价 * 需求量总和
                unit_price = float(ws.cell(row=target_row_idx, column=3).value or 0.0)
                ws.cell(row=target_row_idx, column=tax_price_col_idx, value=unit_price * total_qty_sum)
                
                wb.save(self.filename)
                wb.close()
                
                # 5 & 6. 自动重新刷一遍当前日期标红高亮检测
                self.auto_highlight_today_requirements()
                
                self.status_bar.config(text=f"✅ 成功：产品【{p_id}】在 {target_date} 写入量为 {qty}。全表逻辑公式已同步保存刷新。")
                messagebox.showinfo("更新成功", f"产品编号: {p_id}\n修改列: {target_date}\n新数量: {qty}\n\n该行开票总数与含税总价已自动完成联动乘法计算，数据已完美保留在原表中！")
            else:
                messagebox.showerror("定位失败", "在 Excel 第一列中没有检索到对应的产品编号。")
        except Exception as e:
            messagebox.showerror("文件被占用", f"无法保存到 Excel 电子表！\n请确认您或其他人此时没有用 Excel 打开这个文件。错误原因: {e}")

    def auto_highlight_today_requirements(self):
        """核心功能6：自动检测当前北京时间日期。如表中含当前日期，且该格所需量 > 0，则单独标红"""
        # 获取当前本地时间对应的发货列名 (如当前是3月19日，则匹配 "03-19发货")
        today_column_target = datetime.now().strftime("%m-%d") + "发货"
        
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb["产品发货计划表"]
            headers = [str(cell.value).strip() for cell in ws[1]]
            
            if today_column_target in headers:
                today_col_idx = headers.index(today_column_target) + 1
                # 预设标准的淡红/明红填充样式
                red_fill = PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid")
                clear_fill = PatternFill(fill_type=None) # 清理没需求的格子的颜色
                
                has_highlighted = False
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=today_col_idx)
                    cell_value = cell.value
                    
                    if cell_value and float(cell_value) > 0:
                        cell.fill = red_fill
                        has_highlighted = True
                    else:
                        # 如果没有需求量或者被用户改为了0，自动清除红色，恢复整洁
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == "00FF4D4D":
                            cell.fill = clear_fill
                
                if has_highlighted:
                    wb.save(self.filename)
                    self.status_bar.config(text=f"🚨 变红通知：检测到今日【{today_column_target}】有排产发货所需量，已在 Excel 中单独标红高亮！")
            wb.close()
        except Exception:
            # 启动或保存时静默检测，防止弹窗打扰用户
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = DeliverySystemApp(root)
    root.mainloop()
